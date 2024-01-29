import mysql.connector
from datetime import datetime

class Database:
    def __init__(self, path):
        self.path = path

    def ulanish(self):
        return mysql.connector.connect(
            host = "localhost",
            user = "root",
            password = "Rustam123456b",
            database = "uquv_markaz"
        )

    def ishlatish(self, sql, fetchall=False, fetchone=False, commit=False):
        db = self.ulanish()
        cursor = self.db.cursor()
        data = None
        if fetchall:
            cursor.execute(sql)
            data = cursor.fetchall()
        elif fetchone:
            cursor.execute(sql)
            data = cursor.fetchone()
        elif commit:
            cursor.execute(sql)
            db.commit()

        db.close()
        return data

    def SqlToJadval(self, nomi):
        sql = f"SELECT * FROM {nomi}"
        data = self.ishlatish(sql, fetchall=True)
        excel_table = Excel()
        excel_table.write_table(data)

    def kurs_nomi(self, id):
        sql = f"SELECT kurs_nomi FROM kurslar WHERE id={id}"
        return self.ishlatish(sql, fetchone=True)

    def kurs_nomlari(self):
        sql = f"SELECT kurs_nomi FROM kurslar"
        return self.ishlatish(sql, fetchone=True)

    def uquvchi_id(self,ism, familiya):
        sql = f"""
                SELECT id 
                FROM uquvchilar
                WHERE ism = '{ism}' AND familiya = '{familiya}'
                """
        return self.ishlatish(sql, fetchone=True)

    def kurs_id(self, kurs_nomi):
        sql = f"""
                SELECT id 
                FROM kurslar
                WHERE kurs_nomi = '{kurs_nomi}'
        """
        return self.ishlatish(sql, fetchone=True)

    def insert_tulov(self, ism, familiya, tulov, kurs_nomi):
        UquvchiId = self.uquvchi_id(ism, familiya)
        KursId = self.kurs_id(kurs_nomi)
        sana = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sql = f"""
            INSERT INTO tulovlar 
            VALUES
            ({UquvchiId[0]}, {tulov}, {KursId[0]}, '{sana}')
        """
        self.ishlatish(sql, commit=True)

    def uqituvchi_id(self, ism, familiya):
        sql = f"SELECT id FROM uqituvchilar WHERE ism='{ism}' AND familiya='{familiya}'"
        return self.ishlatish(sql, fetchone=True)

    def id_kurs_nomlari(self, id):
        sql = f"SELECT kurs_nomi FROM kurslar WHERE id = {id}"
        return self.ishlatish(sql, fetchone=True)

    def id_kurslar(self, id):
        sql = f"SELECT kurs_id FROM uqituvchi_kurslari WHERE  uqituvchi_id = {id}"
        return self.ishlatish(sql, fetchone=True)

    def uqituvchi_kurslari(self, ism, familiya):
        UqituvchiId = self.uqituvchi_id(ism, familiya)
        IdKurslar = self.id_kurslar(UqituvchiId[0])
        return self.id_kurs_nomlari(IdKurslar[0])

mybase = Database()
mybase.SqlToJadval('uqituvchilar')

from openpyxl import Workbook, load_workbook

class Excel:
    def __init__(self, name=None, list_nomi=None):
        if name:
            self.file = load_workbook(name)
            self.jadval = self.file[list_nomi]
        else:
            self.file = Workbook()
            self.jadval = self.file.active

    def write_table(self, data):
        self.jadval.append(["№", "ism", "familiya", "T_sana", "Tel"])
        for i in data:
            self.jadval.append(i)
        self.saqlash("uquvchilar.xlsx")

    def saqlash(self, manzil=None):
        self.file.save(manzil)

# print(mybase.uqituvchi_kurslari("Feruza", "Adambayeva"))
# mybase.insert_tulov("Abror", "Habibullayev", 200000, "Kompyuter savodxonligi")
# kurs_nomi = mybase.kurs_nomi(3)
# print(kurs_nomi)
# kurs_nomlari = mybase.kurs_nomlari()
# for i in kurs_nomlari:
#     print(i)
