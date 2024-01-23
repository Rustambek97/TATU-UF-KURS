from openpyxl import Workbook, load_workbook

class Excel:
    def __init__(self, fayl_nomi=None, sheet=None):
        if fayl_nomi == None:
            self.name = fayl_nomi
            self.fayl = Workbook()
            self.jadval = self.fayl.active
        else:
            self.name = fayl_nomi
            self.fayl = load_workbook(f"{fayl_nomi}")
            self.jadval = self.fayl[f"{sheet}"]
        
    def update(self, address, qiymat):
        self.jadval[address] = qiymat
    
    def add(self, *s):
        self.jadval.append(list(s))

    def qushish(self, a, b, c):
        self.jadval[c] = self.jadval[a].value + self.jadval[b].value

    def del_cell(self, a):
        self.jadval[a] = None

    def __repr__(self):
        return f"Fayl nomi: {self.name}"

    def __call__(self, a):
        return self.jadval[a].value
    
    def read_table(self):
        #bu tsiklda barcha qator kataklaridagi ma'lumotlar massiv ko'rinishida chiqadi
        #values_only faqat qiymat kiritilgan kataklarni olish vazifasini bajaradi
        for row in self.jadval.iter_rows(min_row=1, max_row=self.jadval.max_row, values_only=True):
            print(row)
    
    def saqlash(self, s):
        self.jadval.save(s)
        
obj1 = Excel()
obj2 = Excel("python_kurs.xlsx", "Лист1")