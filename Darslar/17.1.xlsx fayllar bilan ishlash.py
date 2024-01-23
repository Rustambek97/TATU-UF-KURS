from openpyxl import *
class Excel:
    def __init__(self, name):
        self.name=name
        self.wb=Workbook()
        self.ws=self.wb.active

    def write_cell(self, address, s):
        self.ws[address] = s

    def add(self, *s):
        self.ws.append(list(s))

    def qushish(self, a, b, c):
        self.ws[c] = self.ws[a].value + self.ws[b].value

    def del_cell(self, a):
        self.ws[a] = None

    def __repr__(self):
        return f"Fayl nomi: {self.name}"

    def __call__(self, a):
        return self.ws[a].value

    def saqlash(self, s):
        self.wb.save(s)

